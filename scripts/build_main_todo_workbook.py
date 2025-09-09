#!/usr/bin/env python3
"""Generate main_to-do.xlsx from individual to-do CSV files with intra-workbook hyperlinks.

Sheets created (if source CSV exists under to-do/):
  - hindsight  (master)
  - security
  - refactors
  - ux
  - capture
  - search
  - config
  - release
  - completed

In the master sheet a column 'Link' is appended. Each row with a Key present in one of the
other domain sheets (matching that sheet's MainKey or Key) will contain a HYPERLINK formula
pointing to the corresponding row on that sheet. Google Sheets and Excel both honor this.

Requirements for compatibility:
  - Simple sheet names (lowercase, no spaces) to avoid quoting issues.
  - Hyperlinks use internal reference pattern: #sheet!A<row>
  - No merged cells or exotic formatting.

Run: python scripts/build_main_todo_workbook.py
"""
from __future__ import annotations
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional
import pandas as pd
from io import StringIO
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parent.parent
TODO_DIR = ROOT / "to-do"
OUTPUT = TODO_DIR / "main_to-do.xlsx"

# Map target sheet names to source CSV filenames
SOURCES = {
    "hindsight": "hindsight_to-do.csv",
    "security": "security_to-do.csv",
    "refactors": "refactors_to-do.csv",
    "ux": "ux_to-do.csv",
    "capture": "capture_to-do.csv",
    "search": "search_to-do.csv",
    "config": "config_to-do.csv",
    "release": "release_to-do.csv",
    "completed": "completed_to-do.csv",
}

MASTER_SHEET = "hindsight"
MASTER_KEY_COL = "Key"
DETAIL_KEY_COLUMNS = ["MainKey", "Key"]  # Order tried when searching in detail sheets
LINK_COL_NAME = "Link"


def read_csv_dataframe(path: Path) -> pd.DataFrame:
    """Read a CSV into DataFrame preserving column order."""
    text = path.read_text(encoding="utf-8").splitlines()
    # Filter out comment lines starting with // if any
    rows = [r for r in text if not r.startswith("//")] if text else []
    if not rows:
        return pd.DataFrame()
    # Use csv.Sniffer? Assume well formed.
    return pd.read_csv(StringIO("\n".join(rows)))


def load_dataframes() -> Dict[str, pd.DataFrame]:
    dfs: Dict[str, pd.DataFrame] = {}
    for sheet, filename in SOURCES.items():
        path = TODO_DIR / filename
        if path.exists():
            try:
                df = read_csv_dataframe(path)
            except Exception as e:  # noqa: BLE001
                print(f"WARN: failed reading {filename}: {e}")
                continue
            if not df.empty:
                dfs[sheet] = df
    return dfs


def write_initial_workbook(dfs: Dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as xw:
        for sheet, df in dfs.items():
            # Ensure no duplicate sheet names
            safe_sheet = sheet[:31]
            df.to_excel(xw, sheet_name=safe_sheet, index=False)


def build_key_index(dfs: Dict[str, pd.DataFrame]) -> Dict[str, tuple[str, int]]:
    """Return mapping of master key -> (sheet, row_number_in_sheet). Row number is 1-based in Excel."""
    index: Dict[str, tuple[str, int]] = {}
    for sheet, df in dfs.items():
        if sheet == MASTER_SHEET:
            continue
        for key_col in DETAIL_KEY_COLUMNS:
            if key_col not in df.columns:
                continue
            for logical_row_idx, value in enumerate(df[key_col].tolist()):
                if isinstance(value, str) and value:
                    # Excel row = header (1) + logical_row_idx + 1
                    excel_row = logical_row_idx + 2
                    # Only record if not already mapped (first sheet wins)
                    index.setdefault(value, (sheet, excel_row))
            break  # Stop after first existing key column
    return index


def append_hyperlinks(dfs: Dict[str, pd.DataFrame]) -> None:
    wb = load_workbook(OUTPUT)
    if MASTER_SHEET not in wb.sheetnames:
        raise RuntimeError("Master sheet missing in workbook")
    ws: Worksheet = wb[MASTER_SHEET]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if LINK_COL_NAME not in header:
        ws.cell(row=1, column=len(header) + 1, value=LINK_COL_NAME)
        header.append(LINK_COL_NAME)
    link_col_index = header.index(LINK_COL_NAME) + 1

    key_index = build_key_index(dfs)

    # Determine master key column index
    try:
        master_key_col_index = header.index(MASTER_KEY_COL) + 1
    except ValueError:
        raise RuntimeError("Master Key column not found in master sheet header")

    for row in range(2, ws.max_row + 1):
        key_val = ws.cell(row=row, column=master_key_col_index).value
        if not key_val:
            continue
        match = key_index.get(str(key_val))
        if match:
            sheet, target_row = match
            # Hyperlink formula (both Excel & Google Sheets)
            formula = f"=HYPERLINK(\"#'{sheet}'!A{target_row}\",\"{sheet}\")"
            ws.cell(row=row, column=link_col_index, value=formula)

    # Basic formatting: autosize columns (approximate) & freeze header
    for col_idx, _ in enumerate(header, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            val = cell.value
            if val is None:
                continue
            l = len(str(val))
            if l > max_len:
                max_len = l
        ws.column_dimensions[col_letter].width = min(60, max_len + 2)

    ws.freeze_panes = "A2"
    wb.save(OUTPUT)
    print(f"Wrote workbook: {OUTPUT}")


def main() -> None:
    dfs = load_dataframes()
    if MASTER_SHEET not in dfs:
        raise SystemExit("Master (hindsight) CSV not found or empty")
    write_initial_workbook(dfs)
    append_hyperlinks(dfs)


if __name__ == "__main__":  # pragma: no cover
    main()
